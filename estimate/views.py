from flask import (Blueprint, render_template, redirect, url_for, 
                    current_app, send_from_directory, request, g)
from apps.app import db
# from apps.crud.models import User
from apps.estimate.models import IncomeState, version, exRate, currentVersion, salesEstimate, specialThings, eliminate
# import uuid
# path를 import한다.
from pathlib import Path
from apps.estimate.forms import UploadImageForm, manageForm, descForm
from flask_login import current_user, login_required
import pandas as pd
import openpyxl
from datetime import datetime
from sqlalchemy import create_engine, select, delete, and_, insert, func
# from sqlalchemy.sql.expression import func
# from sqlalchemy.orm import sessionmaker
from apps.estimate import varGroup, mkHtml, queries
import copy
# from ast import literal_eval

g_year = '2022'
g_version = '2022-10-1st'
g_comp = '1000'
g_mm = '01'
g_tomm = '01'
engine = create_engine(varGroup.mysql_url)

# est를 블루프린트화한다.
est = Blueprint("est", __name__, template_folder="templates")

# 데이터 프레임을 받아 리스트로 전달(주요리스트를 함수화 함)
def dfToList1(df):
    df2 = df.iloc[:, 3:-2].drop(columns=['company','curr'])
    df2.set_index(keys='mm',inplace=True)
    sr = df2.sum(axis=0)
    sr.name = 'total'
    df_t = pd.DataFrame(sr).T
    df2 = pd.concat([df2,df_t],axis=0)
    df_base = pd.DataFrame(df2['Sales']).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
    df_rate = pd.DataFrame()
    for i in range(len(varGroup.targetGroup)):
        df_target = pd.DataFrame(df2[varGroup.targetGroup[i]]).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
        df3 = df_target.div(df_base).mul(100).round(1)
        df4 = df3.set_axis(labels=[varGroup.tGroupName[i]],axis=1)
        df_rate = pd.concat([df_rate, df4],axis=1)
    df3 = pd.concat([df2,df_rate],axis=1)
    df_lv = pd.DataFrame(varGroup.costLabel, index=["lv","명칭","표현"])
    df4 = pd.concat([df_lv,df3],axis=0)
    df5 = df4[varGroup.costElementGroup2].T
    df5.reset_index(inplace=True)
    # idx = df4.pop('index').tolist()
    try:
        del([df,df2,df3,df4,df_rate,df_base,df_t])
    except:
        pass
    return df5

def dfToList2(df):
    df2 = df
    df2.set_index(keys='mm',inplace=True)
    sr = df2.sum(axis=0)
    sr.name = 'total'
    df_t = pd.DataFrame(sr).T
    df2 = pd.concat([df2,df_t],axis=0)
    df_base = pd.DataFrame(df2['Sales']).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
    df_rate = pd.DataFrame()
    for i in range(len(varGroup.targetGroup)):
        df_target = pd.DataFrame(df2[varGroup.targetGroup[i]]).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
        df3 = df_target.div(df_base).mul(100).round(1)
        df4 = df3.set_axis(labels=[varGroup.tGroupName[i]],axis=1)
        df_rate = pd.concat([df_rate, df4],axis=1)
    df3 = pd.concat([df2,df_rate],axis=1)
    df_lv = pd.DataFrame(varGroup.costLabel, index=["lv","명칭","표현"])
    df4 = pd.concat([df_lv,df3],axis=0)
    df5 = df4[varGroup.costElementGroup2].T
    df5.reset_index(inplace=True)
    # idx = df4.pop('index').tolist()
    try:
        del([df,df2,df3,df4,df_rate,df_base,df_t])
    except:
        pass
    return df5



@est.before_request
def authonticate():
    global g_year, g_version, g_comp, g_mm, g_tomm
    sql1 = select(currentVersion)
    with engine.connect() as conn:
        result = conn.execute(sql1).one()
    list_result = list(result)
    print("여기가 첫번째 요청", list_result)
    g_year = list_result[2]
    g_version = list_result[1]
    g_comp = list_result[5]
    g_mm = list_result[3]
    g_tomm = list_result[4]
    # print(g_year, g_version, g_comp)


# dt 애플리케이션을 사용하여 앤드 포인트를 작성한다.
@est.route("/")
def index():
    global g_year, g_version, g_comp, g_mm, g_tomm
    form1 = descForm()
    if g_version[-3:] == '(P)' or g_version[-3:] == '(A)':
        g_mm = '01'
        g_tomm = '12'
    else:
        pass
    l_version = g_version[:4]+'-00-(A)'
    l_version2 = g_version[:4]+'-00-(P)'
    print(g_version[-3:], "확인부탁해요.", l_version, l_version2)
    # 현재 버전 (추정, 실적, 계획)을 조회하는 부분
    sql1 = queries.sum_IS_normal.where(and_(IncomeState.year == g_year, IncomeState.company==g_comp, IncomeState.version == g_version, IncomeState.curr != 'EXW',IncomeState.mm>=g_mm,IncomeState.mm<=g_tomm)).order_by(IncomeState.mm)
    # 이외 부분은 실적(추정기간 앞), 이후는 계획을 넣고 마무리
    sql_act = queries.sum_IS_normal.where(and_(IncomeState.year == g_year, IncomeState.company==g_comp, IncomeState.version == l_version, IncomeState.curr != 'EXW',IncomeState.mm<g_mm)).order_by(IncomeState.mm)
    sql_plan = queries.sum_IS_normal.where(and_(IncomeState.year == g_year, IncomeState.company==g_comp, IncomeState.version == l_version2, IncomeState.curr != 'EXW',IncomeState.mm>g_tomm)).order_by(IncomeState.mm)
    manageList = specialThings.query.filter_by(year=g_year).all()
    
    with engine.connect() as conn:
        result_est = conn.execute(sql1)
        result_act = conn.execute(sql_act)
        result_plan = conn.execute(sql_plan)

    df_est = pd.DataFrame(result_est)
    df_act = pd.DataFrame(result_act)
    df_plan = pd.DataFrame(result_plan)
    print(df_plan)
    if list(df_est) == []:
        finalHtml = "해당월에 추정이 없습니다."
        # pass
    else:
        if list(df_act) == []:
            df = df_est
        else:
            df = pd.concat([df_act, df_est], axis=0)
        if list(df_plan) == []:
            pass
        else:
            df = pd.concat([df, df_plan], axis=0)
        
        df5 = dfToList2(df)
        header_is = df5.columns.values.tolist()
        IncomeStatement = df5.values.tolist()
        finalHtml = mkHtml.rootHtml("예측",header_is,IncomeStatement)
    try:
        del([df,df5,df_act,df_act,df_plan])
    except:
        pass    
    return render_template("estimate/index.html", string2=finalHtml, form_desc = form1, list1 = manageList)

#입력된 화일을 다운 로드 받아볼 부분
@est.route("/images/<path:filename>")
def image_file(filename):
    return send_from_directory(current_app.config["UPLOAD_FOLDER"], filename)


# 정보를 입력하기 위한 부분
# 법인별 손익 계산서를 입력하기 위한 부분
@est.route("/upload", methods=["GET","POST"])
# 로그인 필수로 한다.
@login_required
def upload_image():
    form=UploadImageForm()
    if form.validate_on_submit():
        # 업로드된 이미지 파일을 취득한다.
        file=form.image.data
        # 파일의 파일명과 확장자를 취득하고, 파일명을 uuid로 변화한다.
        ext = Path(file.filename).suffix
        co_code = eval(request.form['ccode'])
        image_uuid_file_name = str(co_code[0]) + "_" + str(request.form['ver']) + ext
        # 이미지를 저장한다.
        image_path = Path(current_app.config["UPLOAD_FOLDER"], image_uuid_file_name)
        file.save(image_path)
        # print("-----------------------------",co_code[0],co_code[1],request.form['year1'],request.form['ver'] )

        # 손익계산서에 내용을 업데이트 한다. 
        file1 = openpyxl.load_workbook(image_path)
        sheet1 = file1.active
        if sheet1['B4'].value != '1. 매출액':
            print("excel 양식을 확인하십시요.",sheet1['B4'].value)
        range1 = sheet1['B3':'P137']
        list1 = []
        for row in range1:
            list_rows = []
            for col in row:
                list_rows.append(col.value)
            list1.append(list_rows)
        list1[0] = ['구분', '삭제', '01', '02',\
         '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '00']
        df1 = pd.DataFrame(list1)
        df2 = df1.iloc[[0,1,6,8,9,10,11,13,15,16,17,18,19,20,21,22,23,24,25,26,27,30,\
            32,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,\
            59,60,61,62,63,65,67,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,86,87,88,\
            89,91,92,93,94,95,96,97,98,99,100,101,102,103,104,106,107,108,109,110,111,112,\
            113,114,115,117,118,119,120,121,122,123,124,125,127,128,129,131,133],\
            [0,2,3,4,5,6,7,8,9,10,11,12,13]].T
        df2.rename(columns=df2.iloc[0],inplace=True)
        df2 = df2.drop(df2.index[0])
        df2["year"]=request.form['year1']
        df2["version"]=request.form['ver']
        df2["company"]=co_code[0]
        df2["curr"]=co_code[1]
        df2["user_id"]=current_user.id
        # 안넣을 경우 어떻게 되는지 확인할 것
        df2["create_at"]=datetime.now
        df2.columns = ['mm','Sales','CGS','CGM','Freight','Insure_ex','GP','SalesExpences',\
            'Agent_ex','Agent_inland','Moving_ex','Moving_inland','Insure_PL','Comp_ex',\
            'Comp_inland','Packing_ex','Packing_inland','Royalty_inland','Royalty_ex',\
            'Advertisement_ex','Advertisement_inland','OperatingProfit1','AdminExpences',\
            'SalaryBOD','Salary','Pension','EmployBenefit','Travel','Telecommunication',\
            'Utilities','TaxesAndDues','Rent','Depreciation','Amortization','Maintanance',\
            'Insurance','Entertainment','SampleFee','TransportFee','Commission','Service',\
            'BedDept','Consumable','Training','BooksAndPrinting','Vehicle','Conference',\
            'Miscellaneous','Overseas','Warehouse','ReversalBadDebt','ResearchDevelopment',\
            'LeaseAmortization','OperatingProfit2','NonOperatingRevenue','FinancialRevenue',\
            'InterestRevenue','TradingSecuritiesInterest','FxRevenue','FxRevenue_Operating',\
            'FxIncone_NonOp','DerivativesRevenue','TradingSecuritiesRevenue',\
            'OtherNonOPRevenue','DividentRevenue','MiscellaneousProfit','Gain_disposal_assets',\
            'Gain_OtherNonOP','ReversalNRV','ReversalOtherAllowance','NonOperatingEnpences',\
            'FinancialCost','Interest','BondInterest','LossDisposalAR','FxLoss','FxLoss_Operating',\
            'FxLoss_NonOP','DerivativesLoss_tot','DerivativesLoss','TradingSecuritiesLoss',\
            'OtherNonOPLoss','Donation','Loss_disposal_assets','Loss_disposal_inventory',\
            'Loss_OtherNonOP','MiscellaneousLoss','OtherAllowance','OrdinaryIncome1',\
            'Gain_disposal_assets_NonOP','DebtForgiveness','LeaseInterestIncome','OtherSpecialGain',\
            'Loss_disposal_assets_NonOP','DebtForgivenessAmortization','LeaseInterestCost',\
            'OtherSpecialLoss','ValuationLoss','OrdinaryIncome2','FxValuation','FxValuationGain_OP',\
            'FxValuationGain_NonOP','FxValuationLoss_OP','FxValuationLoss_NonOP','DerivativesVauation',\
            'DerivativesVauationGain','DerivativesVauationLoss','OrdinaryIncome3','SecuritiesGain',\
            'SecuritiesLoss','OrdinaryIncome4','Tax','NetIncome','year','version','company','curr',\
            'user_id','create_at']
        
        df2=df2.loc[:,\
            ['version','year','mm','company','curr','Sales','CGS','CGM','Freight','Insure_ex','GP','SalesExpences',\
            'Agent_ex','Agent_inland','Moving_ex','Moving_inland','Insure_PL','Comp_ex',\
            'Comp_inland','Packing_ex','Packing_inland','Royalty_inland','Royalty_ex',\
            'Advertisement_ex','Advertisement_inland','OperatingProfit1','AdminExpences',\
            'SalaryBOD','Salary','Pension','EmployBenefit','Travel','Telecommunication',\
            'Utilities','TaxesAndDues','Rent','Depreciation','Amortization','Maintanance',\
            'Insurance','Entertainment','SampleFee','TransportFee','Commission','Service',\
            'BedDept','Consumable','Training','BooksAndPrinting','Vehicle','Conference',\
            'Miscellaneous','Overseas','Warehouse','ReversalBadDebt','ResearchDevelopment',\
            'LeaseAmortization','OperatingProfit2','NonOperatingRevenue','FinancialRevenue',\
            'InterestRevenue','TradingSecuritiesInterest','FxRevenue','FxRevenue_Operating',\
            'FxIncone_NonOp','DerivativesRevenue','TradingSecuritiesRevenue',\
            'OtherNonOPRevenue','DividentRevenue','MiscellaneousProfit','Gain_disposal_assets',\
            'Gain_OtherNonOP','ReversalNRV','ReversalOtherAllowance','NonOperatingEnpences',\
            'FinancialCost','Interest','BondInterest','LossDisposalAR','FxLoss','FxLoss_Operating',\
            'FxLoss_NonOP','DerivativesLoss_tot','DerivativesLoss','TradingSecuritiesLoss',\
            'OtherNonOPLoss','Donation','Loss_disposal_assets','Loss_disposal_inventory',\
            'Loss_OtherNonOP','MiscellaneousLoss','OtherAllowance','OrdinaryIncome1',\
            'Gain_disposal_assets_NonOP','DebtForgiveness','LeaseInterestIncome','OtherSpecialGain',\
            'Loss_disposal_assets_NonOP','DebtForgivenessAmortization','LeaseInterestCost',\
            'OtherSpecialLoss','ValuationLoss','OrdinaryIncome2','FxValuation','FxValuationGain_OP',\
            'FxValuationGain_NonOP','FxValuationLoss_OP','FxValuationLoss_NonOP','DerivativesVauation',\
            'DerivativesVauationGain','DerivativesVauationLoss','OrdinaryIncome3','SecuritiesGain',\
            'SecuritiesLoss','OrdinaryIncome4','Tax','NetIncome']]
        df3=df2.reset_index(drop=True).fillna(0)
        sql1 = delete(IncomeState).where(and_(IncomeState.version == request.form['ver'], IncomeState.company == co_code[0]))
        with engine.connect() as conn:
            result = conn.execute(sql1)
            conn.commit()
        df3.to_sql(name='test_excel2', con=engine, if_exists='append', index=False)
        try:
            del([df,df2,df3])
        except:
            pass
        
        return redirect(url_for("est.index"))

    find_version = select(currentVersion.version).group_by(currentVersion.version)
    find_year = select(currentVersion.year).group_by(currentVersion.year)

    with engine.connect() as conn:
        ver1 = list(conn.execute(find_version))
        YYYY = list(conn.execute(find_year))

    com_code = varGroup.company_code

    return render_template("estimate/upload.html", form=form, ver1 = ver1, YYYY = YYYY, com_code = com_code)

# 환율 입력
@est.route("/exchage_rate", methods=["GET","POST"])
@login_required
def exchangeRate():
    # return render_template("estimate/exchange_rate.html", string2 = "underconstruction")

    form=UploadImageForm()
    if form.validate_on_submit():
        # 업로드된 이미지 파일을 취득한다.
        file=form.image.data
        # 파일의 파일명과 확장자를 취득하고, 파일명을 uuid로 변화한다.
        ext = Path(file.filename).suffix
        file_name = str(request.form['ver'])+ "환율" + ext
        # 이미지를 저장한다.
        image_path = Path(current_app.config["UPLOAD_FOLDER"], file_name)
        # image_path = Path(current_app.config["UPLOAD_FOLDER"], image_uuid_file_name)
        file.save(image_path)

        # 손익계산서에 내용을 업데이트 한다. 
        file1 = openpyxl.load_workbook(image_path)
        sheet1 = file1.active
        if sheet1['A1'].value != '환율':
            print("excel 양식을 확인하십시요.",sheet1['B4'].value)
        range1 = sheet1['A1':'M17']
        list1 = []
        for row in range1:
            list_rows = []
            for col in row:
                list_rows.append(col.value)
            list1.append(list_rows)
        list1[0] = ['Currency','01','02','03','04','05','06','07','08','09','10','11','12']
        df1 = pd.DataFrame(list1)
        
        df1.rename(columns=df1.iloc[0],inplace=True)
        df2 = df1.drop(df1.index[0])
        df2 = df2.melt(id_vars='Currency', value_vars=['01','02','03','04','05','06','07','08','09','10','11','12'])
        df2["year"]=request.form['year1']
        df2["version"]=request.form['ver']
        df2["user_id"]=current_user.id
        # 안넣을 경우 어떻게 되는지 확인할 것
        df2["create_at"]=datetime.now
             
        df2=df2[['version','year','variable','Currency','value','user_id']]
        df2.columns = ['version','year','mm','curr','rate','user_id']
        df3=df2.reset_index(drop=True).fillna(1)
        sql1 = delete(exRate).where(exRate.version == request.form['ver'])
        with engine.connect() as conn:
            result = conn.execute(sql1)
            conn.commit()
        df3.to_sql(name='ex_rate', con=engine, if_exists='append', index=False)
        try:
            del([df1,df2,df3])
        except:
            pass
        return redirect(url_for("est.exchangeRate"))

    find_version = select(currentVersion.version).group_by(currentVersion.version)
    find_year = select(currentVersion.year).group_by(currentVersion.year)
    checkFX = select(exRate.year, exRate.mm, exRate.curr, exRate.rate).where(exRate.version == g_version)

    with engine.connect() as conn:
        ver1 = list(conn.execute(find_version))
        YYYY = list(conn.execute(find_year))
        checkRate = list(conn.execute(checkFX))

    result = []
    header = ['통화','1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월']

    if checkRate == []:
        pass
    else:
        df = pd.DataFrame(checkRate)
        df2 = df.pivot(index = 'curr', columns = 'mm', values='rate')
        df2.reset_index(level=0, inplace=True)
        result = df2.values.tolist()
        try:
            del([df,df2,df3])
        except:
            pass
        
    return render_template("estimate/upload_fx.html", form=form, YYYY = YYYY, ver1 = ver1, result = result, header = header)

# 판매계획 입력
@est.route("/salesplan", methods=["GET","POST"])
@login_required
def salesPlan():
    global g_version, g_year
    form=UploadImageForm()
    if form.validate_on_submit():
        # 업로드된 이미지 파일을 취득한다.
        file=form.image.data
        # 파일의 파일명과 확장자를 취득하고, 파일명을 uuid로 변화한다.
        ext = Path(file.filename).suffix
        file_name = str(request.form['ver'])+ "판매" + ext
        # 이미지를 저장한다.
        image_path = Path(current_app.config["UPLOAD_FOLDER"], file_name)
        # image_path = Path(current_app.config["UPLOAD_FOLDER"], image_uuid_file_name)
        file.save(image_path)

        # 손익계산서에 내용을 업데이트 한다. 
        file1 = openpyxl.load_workbook(image_path)
        sheet1 = file1.active
        if sheet1['A1'].value != '영업그룹':
            print("excel 양식을 확인하십시요.",sheet1['B4'].value)
        range1 = sheet1['A1':'N118']
        list1 = []
        for row in range1:
            list_rows = []
            for col in row:
                list_rows.append(col.value)
            list1.append(list_rows)
        list1[0] = ['salesDivision','curr','01','02','03','04','05','06','07','08','09','10','11','12']
        df1 = pd.DataFrame(list1)
        
        df1.rename(columns=df1.iloc[0],inplace=True)
        df2 = df1.drop(df1.index[0])
        df2 = df2.melt(id_vars=['salesDivision','curr'], value_vars=['01','02','03','04','05','06','07','08','09','10','11','12'])
        df3 = df2.pivot(index=['salesDivision','variable'],columns='curr',values='value')
        df3.columns = df3.columns.values
        df3.reset_index(level=0, inplace=True)
        df3.reset_index(level=0, inplace=True)
        df3["year"]=request.form['year1']
        df3["version"]=request.form['ver']
        # 안넣을 경우 어떻게 되는지 확인할 것
        df3["created_at"]=datetime.now
        df3.columns = ['mm','salesDivision','amount_EA','amount_sales','amount_weight','year','version','created_at']
        df3=df3[['version','year','mm','salesDivision','amount_EA','amount_weight','amount_sales']]
 
        df4=df3.reset_index(drop=True).fillna(0)
        sql1 = delete(salesEstimate).where(salesEstimate.version == request.form['ver'])
        with engine.connect() as conn:
            result = conn.execute(sql1)
            conn.commit()
        df4.to_sql(name='salesplan', con=engine, if_exists='append', index=False)
        try:
            del([df1,df2,df3,df4])
        except:
            pass
        return redirect(url_for("est.salesPlan"))
    find_version = select(currentVersion.version).group_by(currentVersion.version)
    find_year = select(currentVersion.year).group_by(currentVersion.year)
    searchingSales = select(salesEstimate.mm, salesEstimate.amount_sales, salesEstimate.amount_EA, salesEstimate.amount_weight).where(and_(salesEstimate.version == g_version,salesEstimate.salesDivision == '연결'))

    with engine.connect() as conn:
        ver1 = list(conn.execute(find_version))
        YYYY = list(conn.execute(find_year))
        checkSales = list(conn.execute(searchingSales))

    result = []
    header = ['구분','1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월']

    if checkSales == []:
        pass
    else:
        df = pd.DataFrame(checkSales)
        # print(df)
        # df['wonPerWeight'] = df['amount_sales']/df['amount_weight']*1000
        df_target = pd.DataFrame(df['amount_sales']).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
        df_base = pd.DataFrame(df['amount_weight']).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
        df_r = df_target.div(df_base).mul(1000).astype({'col1':'float64'}).round(0)
        
        # df['wonPerWeight'] = df['amount_sales'].div(df['amount_weight']).mul(100)
        # df.fillna(0)
        df1_1 = pd.concat([df,df_r],axis=1)
        df1_1.columns = ['mm','01.매출액(백만)','02.판매수량(천본)','03.중량(ton)','04.중량단가(원/kg)']
        df2 = df1_1.round(0)
        df3 = df2.melt(id_vars='mm', value_vars=['01.매출액(백만)','02.판매수량(천본)','03.중량(ton)','04.중량단가(원/kg)'])
        df3 = df3.pivot(index = 'variable', columns = 'mm', values='value')
        df3.reset_index(level=0, inplace=True)
        result = df3.round(0).values.tolist()
        # result.insert(0,header)
        try:
            del([df,df2,df3,df1_1,df_target,df_base,df_r])
        except:
            pass

    return render_template("estimate/salesplan.html", form=form, YYYY = YYYY, ver1 = ver1, result = result, header = header)

# 연결조정사항 입력 (기간별추정의 경우는 아직 미적용)
@est.route("/elimination", methods=["GET","POST"])
@login_required
def elimination():
    global g_version, g_year
    form=UploadImageForm()
    if form.validate_on_submit():
        # 업로드된 이미지 파일을 취득한다.
        file=form.image.data
        # 파일의 파일명과 확장자를 취득하고, 파일명을 uuid로 변화한다.
        ext = Path(file.filename).suffix
        file_name = str(request.form['ver'])+ "_기타매출원가_" + ext
        # 이미지를 저장한다.
        image_path = Path(current_app.config["UPLOAD_FOLDER"], file_name)
        # image_path = Path(current_app.config["UPLOAD_FOLDER"], image_uuid_file_name)
        file.save(image_path)

        # 손익계산서에 내용을 업데이트 한다. 
        file1 = openpyxl.load_workbook(image_path,data_only=True)
        sheet1 = file1.active
        if sheet1['A1'].value != '내용':
            print("excel 양식을 확인하십시요.",sheet1['A1'].value)
        # range1 = sheet1['A':'O']
        list1 = []
        for row in sheet1.rows:
            list_rows = []
            for col in row:
                list_rows.append(col.value)
            list1.append(list_rows)

        for list_check in list1[1:]:
            if list_check[1] not in varGroup.costElementGroup:
                # 향후에 에러 메세지를 modal에 띄어주는 화면을 보여주도록 만들것
                print("계정이 안 맞아요", list_check[0],list_check[1])
                return redirect(url_for("est.elimination"))
        header = ['content','account','jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec','tot_month']
        list1[0] = header 
        df1 = pd.DataFrame(list1)
        df1.columns = header
        df2 = df1.drop(df1.index[0])
        # print("===================",df2)
        df2["created_at"]=None
        df2["version"] = request.form['ver']
        df2["source"] = 'file'
        sql1 = delete(eliminate).where(eliminate.version == request.form['ver']).where(eliminate.source == 'file')
        with engine.connect() as conn:
            result = conn.execute(sql1)
            conn.commit()
        df2.to_sql(name='elimination', con=engine, if_exists='append', index=False)
        try:
            del([df1,df2])
        except:
            pass
        return redirect(url_for("est.elimination"))
    
    find_version = select(currentVersion.version).group_by(currentVersion.version)
    find_year = select(currentVersion.year).group_by(currentVersion.year)
    sql1 = queries.search_1.where(eliminate.version == g_version)

    with engine.connect() as conn:
        ver1 = list(conn.execute(find_version))
        YYYY = list(conn.execute(find_year))
        checkCGS = list(conn.execute(sql1))

    result = []
    header = ['id','내용','계정','1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월','total']

    if checkCGS == []:
        pass
    else:
        df = pd.DataFrame(checkCGS)
        # print("요기요", df)
        df.columns = header
        df.loc["합계", '1월':'total'] = df.loc[:,'1월':'total'].sum(axis=0)
        df.loc['합계','내용':'계정'] = ['합계','total']
        result = df.round(0).values.tolist()
        # result.insert(0,header)
        try:
            del([df])
        except:
            pass

    return render_template("estimate/elimination.html", form=form, YYYY = YYYY, ver1 = ver1, result = result, header = header)

# 법인별 합산에서 연결 제거 대상 금액을 산출해서 더해주기 (기간별추정의 경우는 아직 미적용)
@est.route("/cal_eli", methods=["POST"])
@login_required
def cal_eli():
    # if request.method == 'POST'
    global g_version, g_year
    # print('=================elimination=================')
    sql_sales = select(salesEstimate.mm, salesEstimate.amount_sales).where(and_(salesEstimate.version==g_version, salesEstimate.salesDivision=='연결')).order_by(salesEstimate.mm)
    sql_tot = select(IncomeState.mm, IncomeState.Sales, IncomeState.Royalty_inland, IncomeState.Royalty_ex).where(and_(IncomeState.version==g_version, IncomeState.company=='9900')).order_by(IncomeState.mm)
    with engine.connect() as conn:
        consolidated_sales = list(conn.execute(sql_sales))
        tot_sales = list(conn.execute(sql_tot))
    
    if consolidated_sales == [] or tot_sales == []:
        pass
    else:
        a_tot=[]
        for idx,a11 in enumerate(consolidated_sales):
            a_tot.append((a11[0], a11[1]-tot_sales[idx][1],a11[1]-tot_sales[idx][1]+tot_sales[idx][2]+tot_sales[idx][3],-tot_sales[idx][2],-tot_sales[idx][3]))
        header = ['version','content','account','jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
        head1 = [('version',g_version,g_version,g_version,g_version),('content','연결제거_매출','연결제거_매출원가','연결제거_상표사용료','연결제거_상표사용료'),('account','Sales','CGM','Royalty_inland','Royalty_inland')]
        
        for idx,aa in enumerate(head1):
            a_tot.insert(idx,aa)
        df = pd.DataFrame(a_tot).T
        df.columns = header
        df1 = df.drop(df.index[0])
        df1.loc[:,'tot_month'] = df.loc[:,'jan':'dec'].sum(axis=1)
        df1["source"] = 'calculation'
        # print(df1)
        sql1 = delete(eliminate).where(eliminate.version == g_version).where(eliminate.source == 'calculation')
        with engine.connect() as conn:
            conn.execute(sql1)
            conn.commit()
        df1.to_sql(name='elimination', con=engine, if_exists='append', index=False)
        try:
            del([df,df1])
        except:
            pass
    
    return redirect(url_for("est.elimination"))

# 법인별 합산에서 연결 제거 대상 금액을 산출해서 더해주기 (기간별추정의 경우는 아직 미적용)
@est.route("/sendToMain", methods=["POST"])
@login_required
def sendToMain():
    global g_version, g_year
    sql1 = delete(IncomeState).where(IncomeState.version == g_version).where(IncomeState.company == '-100')
    sql_entries = queries.sum_eli_element.where(eliminate.version==g_version)
    with engine.connect() as conn:
        result = list(conn.execute(sql_entries))
        conn.execute(sql1)
        conn.commit()
    if result == []:
        pass
    else:
        df = pd.DataFrame(result).T
        df.columns = df.iloc[0]
        df = df.drop(df.index[0])
        df1 = df.reset_index().rename(columns={'index':'mm'})
        df1.columns.name = ''
        exist_field = list(df1.columns)
        check_acc = varGroup.costElementGroup[2:-2]
        add_col = ['version','year','mm','company','curr']
        for idx, aa in enumerate(add_col):
            check_acc.insert(idx, aa)
        add_col = []
        for indivisual_acc in check_acc:
            if indivisual_acc in exist_field:
                pass
            else:
                add_col.append(indivisual_acc)
        df2 = df1.reindex(columns = df1.columns.tolist() + add_col).fillna(0)[check_acc]
        df3 = pd.concat([df2.loc[:,'version':'curr'],df2.loc[:,'Sales':'NetIncome'].astype(dtype='float64', errors='ignore')],axis=1)
        order_lv1 = varGroup.order_cal_lv1
        order_lv2 = varGroup.order_cal_lv2
        company = '-100'
        df3[['version','year','company','curr']] = [g_version,g_version[0:4],company,'KRW']
        for idx in order_lv1:
            df3[idx] = df3[order_lv1[idx]].sum(axis=1)
        for idx in order_lv2:
            df3[idx] = 0
            for idx2 in order_lv2[idx]:
                if idx2[0] == '+':
                    df3[idx] = df3[idx].add(df3[idx2[1]])
                else:
                    df3[idx] = df3[idx].sub(df3[idx2[1]])

        
        df3.to_sql(name='test_excel2', con=engine, if_exists='append', index=False)
        try:
            del([df,df1,df2,df3])
        except:
            pass
    return redirect(url_for("est.elimination"))
    # 끝나고 나면 잘 됐다는 메세지 띄울 것

# 연결항목지우기(초기화 버튼 작용)
@est.route("/cal_init", methods=["POST"])
@login_required
def cal_init():
    global g_version, g_year
    sql1 = delete(eliminate).where(eliminate.version == g_version)
    with engine.connect() as conn:
        conn.execute(sql1)
        conn.commit()
    return redirect(url_for("est.elimination"))

# specialThings 특별관리사항 입력
@est.route("/writeContents", methods=["GET","POST"])
@login_required
def writeContents():
    inputForm = descForm()
    if inputForm.validate_on_submit():
        try:
            totAMT=float(inputForm.tot_amt.data)
        except:
            totAMT=0
        try:
            janAMT=float(inputForm.jan_amt.data)
        except:
            janAMT=0
        try:
            febAMT=float(inputForm.feb_amt.data)
        except:
            febAMT=0
        try:
            marAMT=float(inputForm.mar_amt.data)
        except:
            marAMT=0
        try:
            aprAMT=float(inputForm.apr_amt.data)
        except:
            aprAMT=0
        try:
            mayAMT=float(inputForm.may_amt.data)
        except:
            mayAMT=0
        try:
            junAMT=float(inputForm.jun_amt.data)
        except:
            junAMT=0
        try:
            julAMT=float(inputForm.jul_amt.data)
        except:
            julAMT=0
        try:
            augAMT=float(inputForm.aug_amt.data)
        except:
            augAMT=0
        try:
            sepAMT=float(inputForm.sep_amt.data)
        except:
            sepAMT=0
        try:
            octAMT=float(inputForm.oct_amt.data)
        except:
            octAMT=0
        try:
            novAMT=float(inputForm.nov_amt.data)
        except:
            novAMT=0
        try:
            decAMT=float(inputForm.dec_amt.data)
        except:
            decAMT=0
        input1 = specialThings(
            year = g_year,
            content = inputForm.content.data,
            description = inputForm.description.data,
            amt_tot = totAMT,
            amt_jan = janAMT,
            amt_feb = febAMT,
            amt_mar = marAMT,
            amt_apr = aprAMT,
            amt_may = mayAMT,
            amt_jun = junAMT,
            amt_jul = julAMT,
            amt_aug = augAMT,
            amt_sep = sepAMT,
            amt_oct = octAMT,
            amt_nov = novAMT,
            amt_dec = decAMT,
        )
        db.session.add(input1)
        db.session.commit()
    return redirect(url_for("est.index"))

# 최종 연결하는 부분 (권한관리필요)
@est.route("/consolidating", methods=["GET"])
@login_required
def consolidating():
    global g_version, g_year
    sql_entries = queries.sum_IS_normal_input.where(IncomeState.version == g_version)
    with engine.connect() as conn:
        result=list(conn.execute(sql_entries))
    if result == []:
        pass
    else:
        df = pd.DataFrame(result)
        sql1 = delete(IncomeState).where(and_(IncomeState.version == g_version,IncomeState.company == '1000'))
        with engine.connect() as conn:
            conn.execute(sql1)
            conn.commit()
        df.to_sql(name='test_excel2', con=engine, if_exists='append', index=False)
    return redirect(url_for("est.index"))


#전체적인 세팅을 위한 부분
# 버전 관리를 위한 트리
@est.route("/manageVersion", methods=["GET","POST"])
# 로그인을 필수로 하고 또한 master 이상 권한을 가진 user만 입력한다. 
@login_required
def manageVersion():
    form = manageForm()
    if current_user.auth not in ["root","master"]:
        return redirect(url_for("est.index"))
    if form.validate_on_submit():
        if form.version.data == '(P)' or form.version.data == '(A)':
            version1 = version(
                year=str(form.year1.data),
                mm='01',
                tomm='12',
                version=str(form.year1.data) + "-" + '00' + "-" + form.version.data,
            )
        else:
            version1 = version(
                year=str(form.year1.data),
                mm=str(form.month1.data),
                tomm=str(form.month2.data),
                version=str(form.year1.data) + "-" + str(form.month1.data) + "-" + form.version.data,
            )
        # if version1.is_duplicate_version():
        #     flash("해당 버전은 기 생성되어 있습니다.")
        #     return redirect(url_for("est.manageVersion"))
        
        # 사용자 정보를 등록한다.
        try:
            db.session.add(version1)
            db.session.commit()
        except:
            print("여기에문제가 생겼네요.")
            # pass
        # GET 파라미터에 next키가 존재하고, 값이 없는 경우는 사용자 열람 페이지로 리다이렉트한다.
        next_ = request.args.get("next")
        if next_ is None or not next_.startswith("/"):
            next_ = url_for("est.index")
        return redirect(next_)
    else:
        print(
            "유효성검사가 문제가 있네요."
        )

    find_version = select(version.version).group_by(version.version)
    find_year = select(version.year).group_by(version.year)

    with engine.connect() as conn:
        ver1 = list(conn.execute(find_version))
        YYYY = list(conn.execute(find_year))
    
    com_code = varGroup.company_code
    
    return render_template("estimate/control.html", form = form, ver1 = ver1, YYYY = YYYY, com_code = com_code)


# 기본화면 로그인시 보일 버전을 세팅 함, 추정 버전을 생성함
@est.route("/manage_control", methods=["POST"])
@login_required
def setControls():
    # g_year = request.form['year1']
    # g_version = request.form['ver']
    # g_comp = eval(request.form['ccode'])[0]
    sql_find = select(version).where(version.version == request.form['ver'])
    with engine.connect() as conn:
        result_current = list(conn.execute(sql_find))
    if result_current == []:
        print("해당버전이 없습니다.")
        return redirect(url_for("est.manageVersion"))
    else:
        print("점검", result_current[0], type(result_current[0]), result_current[0][2])
        sql1 = insert(currentVersion).values(version = result_current[0][4], year = result_current[0][1], selectedMM = result_current[0][2], toMM = result_current[0][3], company = eval(request.form['ccode'])[0])
        sql2 = delete(currentVersion)
        with engine.connect() as conn:
            result = conn.execute(sql2)
            conn.commit()
        with engine.connect() as conn:
            result = conn.execute(sql1)
            conn.commit()
    return redirect(url_for("est.index"))


# 환율 환산을 위한 부분 (권한관리필요)
@est.route("/exchangeFX", methods=["GET"])
@login_required
def executeFX():
    # # 향후에 미입력법인이나, 환율의 입력 여부를 판단하여 Modal을 띄우는 부분을 구체화 할 것
    sql1 = queries.exchange_fx_query
    with engine.connect() as conn:
        result_current = list(conn.execute(sql1.where(and_(IncomeState.version == g_version,IncomeState.curr != 'EXW', IncomeState.company >= '1100', IncomeState.company <='9100'))))
    sql1_del = delete(IncomeState).where(and_(IncomeState.version == g_version, IncomeState.curr == 'EXW'))

    if result_current == []:
        pass
    else:
        with engine.connect() as conn:
            result_del = conn.execute(sql1_del)
            conn.commit()
        df1 = pd.DataFrame(result_current)
        print(df1)
        df1.to_sql(name='test_excel2', con=engine, if_exists='append', index=False)
        del(df1)

    return redirect(url_for("est.index"))


# 단순합산을 하는 부분 (권한관리필요)
@est.route("/sumIS", methods=["GET"])
@login_required
def sumIS():
    # # 향후에 미입력법인이나, 환율의 입력 여부를 판단하여 Modal을 띄우는 부분을 구체화 할 것
    # print("visited here!!_____________")
    sql1 = queries.sum_IS_query
    with engine.connect() as conn:
        result_current = list(conn.execute(sql1.where(and_(IncomeState.version == g_version,IncomeState.curr == 'EXW', IncomeState.company >= '1100', IncomeState.company < '9900'))))
    sql1_del = delete(IncomeState).where(and_(IncomeState.version == g_version, IncomeState.company == '9900'))

    if result_current == []:
        pass
    else:
        with engine.connect() as conn:
            conn.execute(sql1_del)
            conn.commit()
        df1 = pd.DataFrame(result_current)
        print(df1)
        df1.to_sql(name='test_excel2', con=engine, if_exists='append', index=False)
        del(df1)

    return redirect(url_for("est.index"))



# 조회화면을 구현하는 부분 나중에 app으로 독립시킬것
# 계획대비
@est.route("/estimation/viewPlan", methods=["GET"])
@login_required
def viewPlan():
    # form1 = descForm()
    # print(g_year+'-00-(P)','--------------')
    sql1 = select(IncomeState).where(and_(IncomeState.year == g_year, IncomeState.company==g_comp, 
    IncomeState.curr!='EXW', 
    IncomeState.version == (g_year+'-00-(P)'))).group_by()
    
    with engine.connect() as conn:
        result = conn.execute(sql1)

    df = pd.DataFrame(result)
    print("또여기",df)
    if list(df) == []:
        finalHtml = "해당월에 추정이 없습니다."
    else:
        df5 = dfToList1(df)
        header_is = df5.columns.values.tolist()
        IncomeStatement = df5.values.tolist()
        finalHtml = mkHtml.rootHtml("예측",header_is,IncomeStatement)
        try:
            del([df,df5])
        except:
            pass
    return render_template("estimate/viewplan.html",string2 = finalHtml)

# 계획대비 : 실적조회
@est.route("/estimation/viewActual", methods=["GET"])
@login_required
def viewActual():
    form1 = descForm()
    print(g_year+'-00-(A)','--------------')
    sql1 = select(IncomeState).where(and_(IncomeState.year == g_year, IncomeState.company==g_comp,IncomeState.mm <= g_mm, 
    IncomeState.curr!='EXW',
    IncomeState.version == (g_year+'-00-(A)'))).group_by()
    
    with engine.connect() as conn:
        result = conn.execute(sql1)

    df = pd.DataFrame(result)
    print("또여기",df)
    if list(df) == []:
        finalHtml = "해당월에 실적이 없습니다."
    else:
        df5 = dfToList1(df)
        header_is = df5.columns.values.tolist()
        IncomeStatement = df5.values.tolist()
        finalHtml = mkHtml.rootHtml("예측",header_is,IncomeStatement)
        try:
            del([df,df5])
        except:
            pass
    return render_template("estimate/viewplan.html",string2 = finalHtml)

# 계획대비
@est.route("/estimation/vsPlan", methods=["GET","POST"])
@login_required
def vsPlan():
    form=UploadImageForm()
    if request.method == "POST":
        # print("여기 점검중", eval(request.form['ccode'])[0])
        # 실적+추정을 DB에서 불러온다.
        sql1 = select(IncomeState).where(and_(IncomeState.year == g_year, IncomeState.company==eval(request.form['ccode'])[0], IncomeState.mm >= request.form['month1'],
        IncomeState.curr!='EXW', IncomeState.mm <= request.form['month2'] ,IncomeState.version == g_version, IncomeState.curr!='EXW')).group_by()
        with engine.connect() as conn:
            result = conn.execute(sql1)
        df = pd.DataFrame(result)

        if list(df) == []:
            finalHtml = "<H3>해당법인의 추정 또는 실적 값이 없습니다.</H3>"
        else:
            df5 = dfToList1(df)
            sql1 = select(IncomeState).where(and_(IncomeState.year == g_year, IncomeState.company==eval(request.form['ccode'])[0], IncomeState.mm >= request.form['month1'], 
            IncomeState.curr!='EXW',
            IncomeState.mm <= request.form['month2'], IncomeState.version == (g_year+'-00-(P)'))).group_by()
            with engine.connect() as conn:
                result = conn.execute(sql1)
            df = pd.DataFrame(result)
            if list(df) == []:
                finalHtml = "<H3>해당법인의 계획 값이 없습니다.</H3>"
            else:
                df5_plan = dfToList1(df)
                df5['계획'] = df5_plan['total']
                df5['계획대비'] = df5['total'] - df5['계획']
                df_base = pd.DataFrame(df5['total']).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
                df_target = pd.DataFrame(df5['계획대비']).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
                df6 = df_target.div(df_base).mul(100).round(2)  #.astype('str')+'%'
                df5['계획비(%)'] = df6['col1']
                # print("나누기\n", df5, "\n", df6, type(df6['col1']))

            header_is = df5.columns.values.tolist()
            IncomeStatement = df5.values.tolist()
            finalHtml = mkHtml.compareHtml("예측",header_is,IncomeStatement)
        try:
            del([df,df5,df5_plan,df6])
        except:
            pass

    else:
        finalHtml = "<H3>조회하고자 하는 기간 및 법인을 선택하십시요.</H3>"

    find_version = select(version.version).group_by(version.version)
    find_year = select(version.year).group_by(version.year)

    with engine.connect() as conn:
        ver1 = list(conn.execute(find_version))
        YYYY = list(conn.execute(find_year))

    com_code = varGroup.company_code
    month_range = ['01','02','03','04','05','06','07','08','09','10','11','12']
    return render_template("estimate/vsPlan.html", form=form, month_range=month_range, string2=finalHtml, ver1=ver1, YYYY=YYYY, com_code=com_code)

# 전년동기대비
@est.route("/estimation/vsPreviosYear", methods=["GET","POST"])
@login_required
def vsPreviousYear():
    form=UploadImageForm()
    if request.method == "POST":
        # print("여기 점검중", eval(request.form['ccode'])[0])
        # 실적+추정을 DB에서 불러온다.
        sql1 = select(IncomeState).where(and_(IncomeState.year == g_year, IncomeState.company==eval(request.form['ccode'])[0], IncomeState.mm >= request.form['month1'],IncomeState.mm <= request.form['month2'],
        IncomeState.curr!='EXW',
        IncomeState.version == g_version)).group_by()
        with engine.connect() as conn:
            result = conn.execute(sql1)
        df = pd.DataFrame(result)

        if list(df) == []:
            finalHtml = "<H3>해당법인의 추정 또는 실적 값이 없습니다.</H3>"
        else:
            df5 = dfToList1(df)
            sql1 = select(IncomeState).where(and_(IncomeState.year == str(int(g_year)-1), IncomeState.company==eval(request.form['ccode'])[0], IncomeState.mm >= request.form['month1'], IncomeState.mm <= request.form['month2'], 
            IncomeState.curr!='EXW',
            IncomeState.version == (g_year+'-00-(A)'))).group_by()
            with engine.connect() as conn:
                result = conn.execute(sql1)
            df = pd.DataFrame(result)
            if list(df) == []:
                df5['전년동기'] = 0
            else:
                df5_plan = dfToList1(df)
                df5['전년동기'] = df5_plan['total']

            df5['동기대비'] = df5['total'] - df5['전년동기']
            df_base = pd.DataFrame(df5['total']).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
            df_target = pd.DataFrame(df5['동기대비']).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
            df6 = df_target.div(df_base).mul(100).round(2)  #.astype('str')+'%'
            df5['동기비(%)'] = df6['col1']
            # print("나누기\n", df5, "\n", df6, type(df6['col1']))

            header_is = df5.columns.values.tolist()
            IncomeStatement = df5.values.tolist()
            finalHtml = mkHtml.compareHtml("예측",header_is,IncomeStatement)
        try:
            del([df,df5,df5_plan,df6])
        except:
            pass

    else:
        finalHtml = "<H3>조회하고자 하는 기간 및 법인을 선택하십시요.</H3>"

    find_version = select(version.version).group_by(version.version)
    find_year = select(version.year).group_by(version.year)

    with engine.connect() as conn:
        ver1 = list(conn.execute(find_version))
        YYYY = list(conn.execute(find_year))

    com_code = varGroup.company_code
    month_range = ['01','02','03','04','05','06','07','08','09','10','11','12']
    return render_template("estimate/vsPreviousYear.html", form=form, month_range=month_range, string2=finalHtml, ver1=ver1, YYYY=YYYY, com_code=com_code)

# 월별 실적 조회 - 계획 대비
@est.route("/estimation/period", methods=["GET","POST"])
@login_required
def period():
    form=UploadImageForm()
    if request.method == "POST":
        # print("여기 점검중", eval(request.form['ccode'])[0])
        # 실적+추정을 DB에서 불러온다.
        sql1 = select(IncomeState).where(and_(IncomeState.year == g_year, IncomeState.company==eval(request.form['ccode'])[0], IncomeState.mm >= request.form['month1'],
        IncomeState.curr!='EXW',
        IncomeState.mm <= request.form['month2'] ,IncomeState.version == g_version)).group_by()
        with engine.connect() as conn:
            result = conn.execute(sql1)
        df = pd.DataFrame(result)

        if list(df) == []:
            finalHtml = "<H3>해당법인의 추정 또는 실적 값이 없습니다.</H3>"
        else:
            df5 = dfToList1(df)
            sql1 = select(IncomeState).where(and_(IncomeState.year == g_year, IncomeState.company==g_comp, IncomeState.mm >= request.form['month1'], IncomeState.mm <= request.form['month2'],
            IncomeState.curr!='EXW',
            IncomeState.version == (g_year+'-00-(P)'))).group_by()
            with engine.connect() as conn:
                result = conn.execute(sql1)
            df = pd.DataFrame(result)
            if list(df) == []:
                finalHtml = "<H3>해당법인의 계획 값이 없습니다.</H3>"
            else:
                df5_plan = dfToList1(df)
                df5['계획'] = df5_plan['total']
                df5['계획대비'] = df5['total'] - df5['계획']
                df_base = pd.DataFrame(df5['total']).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
                df_target = pd.DataFrame(df5['계획대비']).set_axis(labels=['col1'],axis=1).astype({'col1':'float64'})
                df6 = df_target.div(df_base).mul(100).round(2)  #.astype('str')+'%'
                df5['계획비(%)'] = df6['col1']
                # print("나누기\n", df5, "\n", df6, type(df6['col1']))

            header_is = df5.columns.values.tolist()
            IncomeStatement = df5.values.tolist()
            finalHtml = mkHtml.compareHtml("예측",header_is,IncomeStatement)
        try:
            del([df,df5,df5_plan,df6])
        except:
            pass

    else:
        finalHtml = "<H3>조회하고자 하는 기간 및 법인을 선택하십시요.</H3>"

    find_version = select(version.version).group_by(version.version)
    find_year = select(version.year).group_by(version.year)

    with engine.connect() as conn:
        ver1 = list(conn.execute(find_version))
        YYYY = list(conn.execute(find_year))

    com_code = varGroup.company_code
    month_range = ['01','02','03','04','05','06','07','08','09','10','11','12']
    return render_template("estimate/period.html", form=form, month_range=month_range, string2=finalHtml, ver1=ver1, YYYY=YYYY, com_code=com_code)

# 분기별 실적 조회 화면
@est.route("/estimation/quarterly", methods=["GET","POST"])
@login_required
def quarterly():
    form=UploadImageForm()
    if form.validate_on_submit():
        return redirect(url_for("est.quarterly"))

    find_version = select(version.version).group_by(version.version)
    find_year = select(version.year).group_by(version.year)

    with engine.connect() as conn:
        ver1 = list(conn.execute(find_version))
        YYYY = list(conn.execute(find_year))

    com_code = varGroup.company_code

    return render_template("estimate/quarterly.html", form=form, ver1 = ver1, YYYY = YYYY, com_code = com_code)

# 법인별 실적 조회
@est.route("/estimation/entities", methods=["GET","POST"])
@login_required
def entity():
    form=manageForm()
    if form.validate_on_submit():
        return redirect(url_for("est.entity"))
    

    find_version = select(version.version).group_by(version.version)
    find_year = select(version.year).group_by(version.year)

    with engine.connect() as conn:
        ver1 = list(conn.execute(find_version))
        YYYY = list(conn.execute(find_year))

    com_code = varGroup.company_code

    return render_template("estimate/entity.html", form=form, ver1 = ver1, YYYY = YYYY, com_code = com_code)


